"""
engine.py — config-driven Excel -> branded PDF engine.

This is supplyknowledge2productcat.py generalized: every function that used
to read from the hardcoded PATHS / BRAND dicts now takes a DocumentConfig
(see schemas.py). No document-type-specific logic lives here — client
catalog, BU catalog, and quotation sheet all flow through the same code.
"""

import io
import os
import openpyxl
from PIL import Image as PILImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, Image, Flowable
)
from reportlab.pdfgen import canvas as pdfcanvas

from schemas import DocumentConfig

PAGE_W, PAGE_H = A4
MARGIN = 18 * mm
THUMB_SIZE = 28 * mm
PX_PER_EMU = 1 / 9525


# ─────────────────────────────────────────────
#  EXCEL: HEADER DETECTION (used by the upload step, before a preset
#  is even chosen — this is what powers the column-mapping UI)
# ─────────────────────────────────────────────

def detect_headers(path: str, header_row: int = 2) -> list[str]:
    """Return the non-empty header values found in the given row.

    Deliberately NOT using read_only=True here: read-only mode trusts the
    sheet's stored dimension metadata to know how many columns exist, and
    files exported from Google Sheets/Lark often have that metadata wrong
    (truncated), which silently drops real columns. Fully parsing avoids it —
    same approach read_excel() already uses.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row))
    headers = [str(c.value).strip() for c in row if c.value not in (None, "")]
    wb.close()
    return headers


# ─────────────────────────────────────────────
#  FLOATING IMAGE EXTRACTOR (unchanged logic from the original script,
#  just parameterized on which Excel columns to skip and which rows
#  are "valid" per the config's select column)
# ─────────────────────────────────────────────

def extract_images_by_row(ws, valid_rows: set[int], skip_columns: list[int]) -> dict:
    if not ws._images:
        return {}

    row_image_map = {}
    absolute_images = []

    for img in ws._images:
        anchor = img.anchor
        atype = type(anchor).__name__

        if atype in ("TwoCellAnchor", "OneCellAnchor"):
            if anchor._from.col in skip_columns:
                continue
            excel_row = anchor._from.row + 1
            if excel_row in valid_rows:
                row_image_map[excel_row] = io.BytesIO(img._data())

        elif atype == "AbsoluteAnchor":
            y_px = anchor.pos.y * PX_PER_EMU
            absolute_images.append((y_px, img))

    if absolute_images:
        absolute_images.sort(key=lambda t: t[0])
        remaining_rows = [r for r in sorted(valid_rows) if r not in row_image_map]
        for i, (_, img) in enumerate(absolute_images):
            if i >= len(remaining_rows):
                break
            row_image_map[remaining_rows[i]] = io.BytesIO(img._data())

    return row_image_map


# ─────────────────────────────────────────────
#  EXCEL READER — generic over config.columns
# ─────────────────────────────────────────────

def read_excel(path: str, config: DocumentConfig) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header_cells = next(ws.iter_rows(min_row=config.header_row, max_row=config.header_row))
    headers = [c.value for c in header_cells]
    col = {h: i for i, h in enumerate(headers) if h}

    select_idx = col.get(config.select_column)
    if select_idx is None:
        raise ValueError(
            f"Select column '{config.select_column}' not found. Found: {list(col.keys())}"
        )

    text_columns = [c for c in config.columns if c.type == "text"]
    for c in text_columns:
        if c.source_header and c.source_header not in col:
            raise ValueError(
                f"Column '{c.source_header}' (mapped to '{c.label}') not found. "
                f"Found: {list(col.keys())}"
            )

    has_image_column = any(c.type == "image" for c in config.columns)

    # Which rows are included (select_value match)
    valid_rows = set()
    for row_num, row in enumerate(
        ws.iter_rows(min_row=config.data_start_row, values_only=True),
        start=config.data_start_row,
    ):
        if row[select_idx] is None:
            continue
        if str(row[select_idx]).strip().lower() == config.select_value.lower():
            valid_rows.add(row_num)

    row_images = {}
    if has_image_column:
        row_images = extract_images_by_row(ws, valid_rows, config.image_skip_columns)

    items = []
    for row_num, row in enumerate(
        ws.iter_rows(min_row=config.data_start_row, values_only=True),
        start=config.data_start_row,
    ):
        if row_num not in valid_rows:
            continue

        item = {}
        for c in text_columns:
            idx = col.get(c.source_header)
            val = row[idx] if idx is not None else None
            item[c.key] = str(val).strip() if val not in (None, "") else ""

        if has_image_column:
            item["_image"] = row_images.get(row_num)

        items.append(item)

    wb.close()
    return items


# ─────────────────────────────────────────────
#  BRANDED CANVAS — same as original, reads from config.brand
# ─────────────────────────────────────────────

def make_branded_canvas(config: DocumentConfig, assets_dir: str):
    brand = config.brand
    c_primary = HexColor(brand.primary)
    c_accent = HexColor(brand.accent)
    c_light_bg = HexColor(brand.light_bg)
    c_mid = HexColor(brand.text_mid)

    banner_left = os.path.join(assets_dir, brand.banner_left) if brand.banner_left else None
    banner_right = os.path.join(assets_dir, brand.banner_right) if brand.banner_right else None

    class BrandedCanvas(pdfcanvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_page(total)
                super().showPage()
            super().save()

        def _draw_page(self, total):
            w, h = A4

            self.setFillColor(colors.white)
            self.rect(0, h - 22 * mm, w, 22 * mm, fill=1, stroke=0)

            if banner_left and os.path.exists(banner_left):
                self.drawImage(banner_left, -43, h - 22 * mm, width=w / 2,
                                height=22 * mm, preserveAspectRatio=True, mask="auto")
            if banner_right and os.path.exists(banner_right):
                self.drawImage(banner_right, w * 0.59, h - 22 * mm, width=w / 2,
                                height=22 * mm, preserveAspectRatio=True, mask="auto")
            elif not banner_left:
                self.setFillColor(colors.white)
                self.setFont("Helvetica-Bold", 13)
                self.drawString(MARGIN, h - 13 * mm, brand.company_name)
                self.setFont("Helvetica", 8)
                self.setFillColor(c_accent)
                self.drawString(MARGIN, h - 18.5 * mm, brand.tagline)

            self.setFillColor(c_primary)
            self.rect(0, h - 28.5 * mm, w, 2.5 * mm, fill=1, stroke=0)

            self.setFillColor(c_light_bg)
            self.rect(0, 0, w, 12 * mm, fill=1, stroke=0)
            self.setFillColor(c_accent)
            self.rect(0, 12 * mm, w, 0.8 * mm, fill=1, stroke=0)

            self.setFillColor(c_mid)
            self.setFont("Helvetica", 7.5)
            self.drawString(MARGIN - 5 * mm, 4.5 * mm, f"{brand.company_name} --- {brand.tagline}")
            self.drawRightString(w - MARGIN + 5 * mm, 4.5 * mm, f"Page {self._pageNumber} of {total}")

    return BrandedCanvas


# ─────────────────────────────────────────────
#  THUMBNAIL / PLACEHOLDER HELPER
# ─────────────────────────────────────────────

def make_thumbnail(image_bytes, c_light_bg, c_accent, c_mid, size=THUMB_SIZE):
    if image_bytes:
        try:
            image_bytes.seek(0)
            pil_img = PILImage.open(image_bytes)
            if pil_img.mode in ("RGBA", "P"):
                pil_img = pil_img.convert("RGB")
            px = int(size * 11.8)
            pil_img = pil_img.resize((px, px), PILImage.LANCZOS)
            compressed = io.BytesIO()
            pil_img.save(compressed, format="JPEG", quality=95, optimize=True)
            compressed.seek(0)
            return Image(compressed, width=size, height=size)
        except Exception:
            pass

    class PlaceholderBox(Flowable):
        def __init__(self, s):
            self.width = s
            self.height = s

        def draw(self):
            self.canv.setFillColor(c_light_bg)
            self.canv.setStrokeColor(c_accent)
            self.canv.roundRect(0, 0, self.width, self.height, 4, fill=1, stroke=1)
            self.canv.setFillColor(c_mid)
            self.canv.setFont("Helvetica", 6)
            self.canv.drawCentredString(self.width / 2, self.height / 2 - 3, "No Image")

    return PlaceholderBox(size)


# ─────────────────────────────────────────────
#  TABLE BUILDER — generic over config.columns
# ─────────────────────────────────────────────

def build_table(items: list[dict], config: DocumentConfig, styles: dict,
                 c_primary, c_accent, c_light_bg, c_mid, c_dark, c_white):
    # Decide which optional columns actually show, based on real data
    active_columns = []
    for c in config.columns:
        if c.optional and not any(item.get(c.key) for item in items):
            continue
        active_columns.append(c)

    usable_w = PAGE_W - 2 * MARGIN + 7 * mm
    fixed_total = sum((c.width_mm or 0) * mm for c in active_columns if c.width_mode == "fixed")
    flex_cols = [c for c in active_columns if c.width_mode == "flex"]
    flex_total_weight = sum(c.flex_weight for c in flex_cols) or 1.0
    remaining = usable_w - fixed_total

    col_widths = []
    headers = []
    for c in active_columns:
        if c.type == "image":
            w = THUMB_SIZE + 4 * mm
        elif c.width_mode == "fixed":
            w = c.width_mm * mm
        else:
            w = remaining * (c.flex_weight / flex_total_weight)
        col_widths.append(w)
        headers.append(Paragraph(c.label, styles["col_header"]))

    data = [headers]
    for item in items:
        row = []
        for c in active_columns:
            if c.type == "image":
                row.append(make_thumbnail(item.get("_image"), c_light_bg, c_accent, c_mid))
            else:
                if c.emphasis:
                    style = styles["product_name"]
                elif c.align == "center":
                    style = styles["price_text"]
                else:
                    style = styles["dim_text"]
                row.append(Paragraph(item.get(c.key, ""), style))
        data.append(row)

    has_image = any(c.type == "image" for c in active_columns)
    row_h = (THUMB_SIZE + 4 * mm) if has_image else 10 * mm

    t = Table(data, colWidths=col_widths, rowHeights=[10 * mm] + [row_h] * len(items))
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), c_primary),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("LEFTPADDING", (1, 1), (1, -1), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
        ("LINEAFTER", (0, 0), (-2, -1), 0.5, HexColor("#CCCCCC")),
        ("LINEBEFORE", (0, 0), (0, -1), 3, c_accent),
        ("BOX", (0, 0), (-1, -1), 1, c_primary),
    ]))
    return t


# ─────────────────────────────────────────────
#  FOOTER BLOCKS — generic, driven by config.footer_blocks
# ─────────────────────────────────────────────

def build_footer_blocks(config: DocumentConfig, styles: dict, c_mid, c_dark):
    flowables = []
    for block in config.footer_blocks:
        if block.type == "text" and block.text:
            flowables.append(Spacer(1, 8 * mm))
            flowables.append(Paragraph(block.text, styles["intro"]))
        elif block.type == "signature":
            flowables.append(Spacer(1, 16 * mm))
            n = len(block.labels)
            col_w = (PAGE_W - 2 * MARGIN) / n
            line_row = ["_" * 28 for _ in block.labels]
            label_row = [Paragraph(lbl, styles["dim_text"]) for lbl in block.labels]
            sig_table = Table([line_row, label_row], colWidths=[col_w] * n)
            sig_table.setStyle(TableStyle([
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]))
            flowables.append(sig_table)
    return flowables


# ─────────────────────────────────────────────
#  STYLES
# ─────────────────────────────────────────────

def make_styles(c_dark, c_mid, c_white, c_primary):
    return {
        "product_name": ParagraphStyle("ProductName", fontName="Helvetica-Bold",
                                        fontSize=9, textColor=c_dark, leading=13),
        "dim_text": ParagraphStyle("DimText", fontName="Helvetica",
                                    fontSize=8, textColor=c_mid, leading=11, spaceBefore=3),
        "price_text": ParagraphStyle("PriceText", fontName="Helvetica",
                                      fontSize=8, textColor=c_mid, leading=11, spaceBefore=3,
                                      alignment=TA_CENTER),
        "col_header": ParagraphStyle("ColHeader", fontName="Helvetica-Bold",
                                      fontSize=8.5, textColor=c_white, alignment=TA_CENTER),
        "cat_title": ParagraphStyle("CatTitle", fontName="Helvetica-Bold",
                                     fontSize=18, textColor=c_primary, spaceAfter=12,
                                     leftIndent=-(7 * mm)),
        "intro": ParagraphStyle("Intro", fontName="Helvetica",
                                 fontSize=9, textColor=c_mid, spaceAfter=13,
                                 leftIndent=-(7 * mm)),
    }


# ─────────────────────────────────────────────
#  PDF ASSEMBLER
# ─────────────────────────────────────────────

def generate_pdf(items: list[dict], config: DocumentConfig, output_path: str, assets_dir: str):
    brand = config.brand
    c_primary = HexColor(brand.primary)
    c_accent = HexColor(brand.accent)
    c_light_bg = HexColor(brand.light_bg)
    c_dark = HexColor(brand.text_dark)
    c_mid = HexColor(brand.text_mid)
    c_white = colors.white

    styles = make_styles(c_dark, c_mid, c_white, c_primary)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=30 * mm, bottomMargin=20 * mm,
        title=config.document_title,
        author=brand.company_name,
        compress=1,
    )

    story = [
        Spacer(1, 4 * mm),
        Paragraph(config.document_title, styles["cat_title"]),
        Paragraph(config.intro_text_template.format(count=len(items)), styles["intro"]),
        build_table(items, config, styles, c_primary, c_accent, c_light_bg, c_mid, c_dark, c_white),
    ]
    story.extend(build_footer_blocks(config, styles, c_mid, c_dark))

    canvas_cls = make_branded_canvas(config, assets_dir)
    doc.build(story, canvasmaker=canvas_cls)
    return output_path
