"""
test_engine.py — proves the generalized engine works for all three
current use cases before we build any API/UI on top of it.
"""

import io
import json
import os
import re
import zipfile
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from schemas import DocumentConfig
import engine

HERE = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(HERE, "assets")
OUT = os.path.join(HERE, "output")
os.makedirs(OUT, exist_ok=True)


def make_dummy_image_bytes(color):
    img = PILImage.new("RGB", (80, 80), color)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def build_client_catalog_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["-- title row --"])
    ws.append(["Select", "Product Name", "Dimension (Spec)", "Supply Advantage", "Price Range (THB/kg)"])
    rows = [
        ["Yes", "Frozen Salmon Fillet", "200g x 20 pcs", "Year-round supply, cold chain verified", "180-220"],
        ["Yes", "Chicken Breast", "1kg x 10 pcs", "Local sourcing, 48hr lead time", "120-150"],
        ["No", "Excluded Item", "n/a", "n/a", "n/a"],
        ["Yes", "Tiger Prawns", "500g x 12 pcs", "", "350-400"],
    ]
    for r in rows:
        ws.append(r)

    for row_idx, color in [(3, (200, 80, 80)), (4, (80, 160, 90))]:
        img = XLImage(make_dummy_image_bytes(color))
        img.width, img.height = 80, 80
        ws.add_image(img, f"A{row_idx}")

    wb.save(path)


def build_quotation_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["-- title row --"])
    ws.append(["Include", "Item Name", "Quantity", "Unit Price", "Remarks"])
    rows = [
        ["Yes", "Cardboard Box (M)", "500", "12.50", "Bulk discount applied"],
        ["Yes", "Pallet Wrap", "50", "45.00", ""],
        ["No", "Skip Me", "1", "1", ""],
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)


def run_case(name, xlsx_builder, preset_file):
    print(f"\n=== {name} ===")
    xlsx_path = os.path.join(OUT, f"{name}_input.xlsx")
    xlsx_builder(xlsx_path)

    with open(os.path.join(HERE, "presets", preset_file)) as f:
        config = DocumentConfig(**json.load(f))

    headers = engine.detect_headers(xlsx_path, config.header_row)
    print(f"Detected headers: {headers}")

    items = engine.read_excel(xlsx_path, config)
    print(f"Items included: {len(items)}")
    for it in items:
        print(f"  {it}")

    out_pdf = os.path.join(OUT, f"{name}.pdf")
    engine.generate_pdf(items, config, out_pdf, ASSETS)
    size_kb = os.path.getsize(out_pdf) / 1024
    print(f"PDF generated: {out_pdf} ({size_kb:.1f} KB)")


def test_truncated_dimension_metadata():
    """
    Regression test: files exported from Google Sheets/Lark can have a
    <dimension> tag in the sheet XML that under-reports the real column
    range. detect_headers() must not silently drop columns because of this
    (found via a real Cogistics export — see engine.py detect_headers docstring).
    """
    print("\n=== truncated_dimension_metadata (regression) ===")
    src = os.path.join(OUT, "dim_regression_source.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["title"])
    ws.append(["Select", "Product Name", "Dimension (Spec)", "Supply Advantage", "Price Range (THB/kg)"])
    ws.append(["Yes", "Item A", "10x10", "adv", "100"])
    wb.save(src)

    with zipfile.ZipFile(src, "r") as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode()
    bad_xml = re.sub(r'<dimension ref="[^"]*"/>', '<dimension ref="A1:A3"/>', sheet_xml)

    corrupted = os.path.join(OUT, "dim_regression_corrupted.xlsx")
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(corrupted, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = bad_xml.encode()
            zout.writestr(item, data)

    headers = engine.detect_headers(corrupted, header_row=2)
    print(f"Detected headers: {headers}")
    assert len(headers) == 5, f"Expected 5 headers, got {len(headers)}: {headers}"
    print("PASS")


def test_signature_only_on_last_page():
    """
    Regression test: the quotation preset's signature block must appear
    exactly once, pinned near the bottom of the LAST page — never on
    earlier pages, and never flowing loose wherever the table happens to end.
    """
    print("\n=== signature_only_on_last_page (regression) ===")
    with open(os.path.join(HERE, "presets", "quotation_sheet.json")) as f:
        config = DocumentConfig(**json.load(f))

    xlsx_path = os.path.join(OUT, "sig_regression_input.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["title"])
    ws.append(["Include", "Item Name", "Quantity", "Unit Price", "Remarks"])
    for i in range(40):  # enough rows to force multiple pages
        ws.append(["Yes", f"Item {i+1:02d}", str(10 * (i + 1)), f"{5.5+i}", "remark"])
    wb.save(xlsx_path)

    items = engine.read_excel(xlsx_path, config)
    out_pdf = os.path.join(OUT, "sig_regression_output.pdf")
    engine.generate_pdf(items, config, out_pdf, ASSETS)

    import subprocess
    result = subprocess.run(["pdftotext", "-layout", out_pdf, "-"], capture_output=True, text=True)
    pages = result.stdout.split("\x0c")  # form-feed separates pages in pdftotext output
    pages = [p for p in pages if p.strip()]
    print(f"Generated {len(pages)} page(s)")
    assert len(pages) >= 2, "Test setup should force multiple pages — check row count"

    occurrences = [i for i, p in enumerate(pages) if "Prepared by" in p]
    print(f"'Prepared by' found on page index(es): {occurrences}")
    assert occurrences == [len(pages) - 1], (
        f"Expected signature only on the last page (index {len(pages)-1}), found on {occurrences}"
    )
    print("PASS")


def test_thai_text_renders():
    """
    Regression test: Thai text must render as real embedded text (not boxes,
    not silently dropped). Found via a real Cogistics export with a Thai
    company name — Helvetica has no Thai glyphs and reportlab silently drew
    black boxes instead of raising an error. Extracts text back out of the
    generated PDF with pdftotext to prove the actual Thai characters are
    embedded, not just visually present in a screenshot.
    """
    print("\n=== thai_text_renders (regression) ===")
    xlsx_path = os.path.join(OUT, "thai_regression_input.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["title"])
    ws.append(["Select", "Product Name", "Dimension (Spec)", "Supply Advantage", "Price Range (THB/kg)"])
    thai_name = "บริษัท ออน-กรีน โปรดิวส์ จำกัด"
    ws.append(["Yes", thai_name, "10x10", "adv", "100"])
    wb.save(xlsx_path)

    with open(os.path.join(HERE, "presets", "client_catalog.json")) as f:
        config = DocumentConfig(**json.load(f))

    items = engine.read_excel(xlsx_path, config)
    out_pdf = os.path.join(OUT, "thai_regression_output.pdf")
    engine.generate_pdf(items, config, out_pdf, ASSETS)

    import subprocess
    result = subprocess.run(["pdftotext", "-layout", out_pdf, "-"], capture_output=True, text=True)
    extracted = result.stdout
    # Check each word independently rather than requiring one exact contiguous
    # match: pdftotext -layout interleaves table columns in reading order, so
    # when a long cell value wraps onto two lines, other columns' content from
    # the same row gets extracted in between — a text-extraction-order quirk,
    # not a rendering defect (confirmed correct visually via screenshot).
    words = thai_name.replace("-", " ").split()
    missing = [w for w in words if w not in extracted]
    assert not missing, (
        f"Thai word(s) not found as real text in the generated PDF (rendered "
        f"as boxes or dropped): {missing}. Extracted text was: {extracted!r}"
    )
    print("PASS: Thai text extracted correctly from the generated PDF")


if __name__ == "__main__":
    run_case("client_catalog", build_client_catalog_xlsx, "client_catalog.json")
    run_case("quotation_sheet", build_quotation_xlsx, "quotation_sheet.json")
    test_truncated_dimension_metadata()
    test_signature_only_on_last_page()
    test_thai_text_renders()
    print("\nAll cases passed.")
