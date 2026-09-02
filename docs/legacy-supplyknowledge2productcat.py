"""
Excel → Branded PDF Product Catalog
=====================================
Supports floating images (as exported from Lark/Google Sheets).
Images are matched to rows by their vertical position in the sheet.

Usage:
    python excel_to_pdf.py

Configuration:
    Edit the BRAND and PATHS sections below.
"""

import io
import os
import sys
import openpyxl
from reportlab.lib.pagesizes import A4
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, Image, HRFlowable
)
from reportlab.pdfgen import canvas as pdfcanvas


# ─────────────────────────────────────────────
#  CONFIGURATION — edit these to match your brand
# ─────────────────────────────────────────────

def resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

BRAND = {
    "company_name": "COGISTICS",
    "tagline":      "The One Stop Cold Chain Solution",
    "primary":      "#1A3C5E",   # Dark navy — header bar, table header
    "accent":       "#7C7C7C",   # Gold — stripe, borders
    "light_bg":     "#F5F7FA",   # Alternating row background
    "text_dark":    "#1C1C1C",
    "text_mid":     "#555555",
    "banner_left":   resource_path("cogistics header banner left.png"),
    "banner_right":  resource_path("cogistics header banner right.png"),  # e.g. "logo.png" — or None for text fallback
}

PATHS = {
    "excel_file":  "productsTEST.xlsx",       # Path to your Excel file
    "output_pdf":  "/Users/tai/Desktop/Cogistics Product Suggestions Catalog/product_catalogFN2.pdf", # Where to save the PDF
    # Column headers — must match your Excel headers exactly (case-sensitive)
    "name_column": "Product Name",
    "dim_column":  "Dimension (Spec)",
    "price_column": "Price Range (THB/kg)",
    "advantage_column": "Supply Advantage",
    "select_column": "Select",
    # Note: no "image_column" needed — images are extracted as floating
}

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────

PAGE_W, PAGE_H = A4
MARGIN      = 18 * mm
THUMB_SIZE  = 28 * mm
PX_PER_EMU  = 1 / 9525

C_PRIMARY  = HexColor(BRAND["primary"])
C_ACCENT   = HexColor(BRAND["accent"])
C_LIGHT_BG = HexColor(BRAND["light_bg"])
C_WHITE    = colors.white
C_DARK     = HexColor(BRAND["text_dark"])
C_MID      = HexColor(BRAND["text_mid"])


# ─────────────────────────────────────────────
#  FLOATING IMAGE EXTRACTOR
#  Reads all floating images from the sheet and
#  maps them to data rows by vertical sort order.
# ─────────────────────────────────────────────

def extract_images_by_row(ws, select_idx=None):
    """
    Returns: { excel_row_number (int) -> image BytesIO }

    Handles all three anchor types produced by Excel / Lark / Google Sheets:
      - TwoCellAnchor  → uses _from.row (0-indexed), so excel_row = _from.row + 1
      - OneCellAnchor  → same, uses _from.row
      - AbsoluteAnchor → uses pos.y in EMU; images sorted top-to-bottom
                         and assigned in order to valid rows

    Only maps images to rows where Select column is "Yes".
    Skips images in column F (index 5) — internal use only.
    """
    if not ws._images:
        print("  → No floating images found in sheet")
        return {}

    PX_PER_EMU = 1 / 9525
    row_image_map = {}
    absolute_images = []

    # Build set of valid (Yes) row numbers
    valid_rows = set()
    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if select_idx is None or str(row[select_idx] or "").strip().lower() == "yes":
            valid_rows.add(row_num)

    for img in ws._images:
        anchor = img.anchor
        atype  = type(anchor).__name__

        if atype in ("TwoCellAnchor", "OneCellAnchor"):
            if anchor._from.col == 5:  # skip col F — internal photos
                continue
            excel_row = anchor._from.row + 1
            if excel_row in valid_rows:
                row_image_map[excel_row] = io.BytesIO(img._data())

        elif atype == "AbsoluteAnchor":
            y_px = anchor.pos.y * PX_PER_EMU
            absolute_images.append((y_px, img))

    # Handle AbsoluteAnchor images by sorting Y → assign to valid rows in order
    if absolute_images:
        absolute_images.sort(key=lambda t: t[0])
        remaining_rows = [r for r in sorted(valid_rows) if r not in row_image_map]
        for i, (_, img) in enumerate(absolute_images):
            if i >= len(remaining_rows):
                break
            row_image_map[remaining_rows[i]] = io.BytesIO(img._data())

    total_valid = len(valid_rows)
    mapped = len(row_image_map)
    print(f"  → {mapped}/{total_valid} images mapped to Yes rows")
    if mapped < total_valid:
        print(f"  ⚠  {total_valid - mapped} Yes row(s) have no image — will show placeholder")

    return row_image_map


# ─────────────────────────────────────────────
#  EXCEL READER
# ─────────────────────────────────────────────

def read_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    col = {h: i for i, h in enumerate(headers) if h}

    name_idx = col.get(PATHS["name_column"])
    dim_idx  = col.get(PATHS["dim_column"])
    price_idx = col.get(PATHS["price_column"])
    advantage_idx = col.get(PATHS["advantage_column"])
    select_idx = col.get(PATHS["select_column"])

    if name_idx is None:
        raise ValueError(f"Column '{PATHS['name_column']}' not found. Found: {list(col.keys())}")
    if dim_idx is None:
        raise ValueError(f"Column '{PATHS['dim_column']}' not found. Found: {list(col.keys())}")

    print(f"Reading: {path}")
    print(f"  Columns found: {list(col.keys())}")

    row_images = extract_images_by_row(ws, select_idx=select_idx)

    products = []
    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if select_idx is None or str(row[select_idx] or "").strip().lower() != "yes":
            continue
        name = row[name_idx] or ""
        dim  = row[dim_idx]  or ""
        if not str(name).strip():
            continue
        products.append({
            "name":  str(name),
            "dim":   str(dim),
            "image": row_images.get(row_num),
            "price": str(row[price_idx]).strip() if (price_idx is not None and row[price_idx]) else "",# BytesIO or None
            "advantage": str(row[advantage_idx]).strip() if (advantage_idx is not None and row[advantage_idx]) else "",
        })

    print(f"  {len(products)} products loaded")
    return products


# ─────────────────────────────────────────────
#  BRANDED CANVAS — header & footer on every page
# ─────────────────────────────────────────────

class BrandedCanvas(pdfcanvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_page(total)
            super().showPage()
        super().save()

    def _draw_page(self, total):
        w, h = A4

        # Header bar
        self.setFillColor(C_WHITE)
        self.rect(0, h - 22*mm, w, 22*mm, fill=1, stroke=0)

        # Left banner
        if BRAND.get("banner_left") and os.path.exists(BRAND["banner_left"]):
            self.drawImage(
                BRAND["banner_left"],
                -43, h - 22*mm,
                width=w/2, height=22*mm,
                preserveAspectRatio=True,
                mask="auto"
            )

        # Right banner
        if BRAND.get("banner_right") and os.path.exists(BRAND["banner_right"]):
            self.drawImage(
                BRAND["banner_right"],
                w*0.59, h - 22*mm,
                width=w/2, height=22*mm,
                preserveAspectRatio=True,
                mask="auto"
            )
        else:
            self.setFillColor(C_WHITE)
            self.setFont("Helvetica-Bold", 13)
            self.drawString(MARGIN, h - 13*mm, BRAND["company_name"])
            self.setFont("Helvetica", 8)
            self.setFillColor(C_ACCENT)
            self.drawString(MARGIN, h - 18.5*mm, BRAND["tagline"])

        # Accent stripe below header
        self.setFillColor(C_PRIMARY)
        self.rect(0, h - 28.5*mm, w, 2.5*mm, fill=1, stroke=0)

        # Footer bar
        self.setFillColor(C_LIGHT_BG)
        self.rect(0, 0, w, 12*mm, fill=1, stroke=0)
        self.setFillColor(C_ACCENT)
        self.rect(0, 12*mm, w, 0.8*mm, fill=1, stroke=0)

        self.setFillColor(C_MID)
        self.setFont("Helvetica", 7.5)
        self.drawString(MARGIN - 5*mm, 4.5*mm, BRAND["company_name"] + " --- " + BRAND["tagline"])
        self.drawRightString(w - MARGIN + 5*mm, 4.5*mm, f"Page {self._pageNumber} of {total}")


# ─────────────────────────────────────────────
#  STYLES
# ─────────────────────────────────────────────

def make_styles():
    return {
        "product_name": ParagraphStyle(
            "ProductName", fontName="Helvetica-Bold",
            fontSize=9, textColor=C_DARK, leading=13,
        ),
        "dim_text": ParagraphStyle(
            "DimText", fontName="Helvetica",
            fontSize=8, textColor=C_MID, leading=11, spaceBefore=3,
        ),
        "price_text": ParagraphStyle(
            "PriceText", fontName="Helvetica",
            fontSize=8, textColor=C_MID, leading=11, spaceBefore=3,
            alignment=TA_CENTER,
        ),
        "col_header": ParagraphStyle(
            "ColHeader", fontName="Helvetica-Bold",
            fontSize=8.5, textColor=C_WHITE, alignment=TA_CENTER,
        ),
        "cat_title": ParagraphStyle(
            "CatTitle", fontName="Helvetica-Bold",
            fontSize=18, textColor=C_PRIMARY, spaceAfter=2,
        ),
        "intro": ParagraphStyle(
            "Intro", fontName="Helvetica",
            fontSize=9, textColor=C_MID, spaceAfter=10,
        ),
    }


# ─────────────────────────────────────────────
#  THUMBNAIL HELPER
# ─────────────────────────────────────────────

def make_thumbnail(image_bytes, size=THUMB_SIZE):
    """Return a ReportLab Image from BytesIO, or a placeholder if None."""
    if image_bytes:
        try:
            image_bytes.seek(0)
            pil_img = PILImage.open(image_bytes)
            
            if pil_img.mode in ("RGBA", "P"):
                pil_img = pil_img.convert("RGB")
            
            px = int(size * 11.8)
            pil_img = pil_img.resize((px, px), PILImage.LANCZOS)
            
            compressed = io.BytesIO()
            pil_img.save(compressed, format="JPEG", quality=95, optimize=True)
            compressed.seek(0)
            return Image(compressed, width=size, height=size)
        except Exception as e:
            print(f"  ⚠  Could not render image: {e}")

    from reportlab.platypus import Flowable
    class PlaceholderBox(Flowable):
        def __init__(self, s):
            self.width = s
            self.height = s
        def draw(self):
            self.canv.setFillColor(C_LIGHT_BG)
            self.canv.setStrokeColor(C_ACCENT)
            self.canv.roundRect(0, 0, self.width, self.height, 4, fill=1, stroke=1)
            self.canv.setFillColor(C_MID)
            self.canv.setFont("Helvetica", 6)
            self.canv.drawCentredString(self.width/2, self.height/2 - 3, "No Image")
    return PlaceholderBox(size)


# ─────────────────────────────────────────────
#  PRODUCT TABLE
# ─────────────────────────────────────────────

def build_product_table(products, styles):
    show_price = any(p["price"] for p in products)
    show_advantage = any(p["advantage"] for p in products)
    
    usable_w  = PAGE_W - 2 * MARGIN + 7*mm
    col_photo = THUMB_SIZE + 4*mm
    col_dim   = 25*mm
    col_price = 30*mm # only if price exists
    
    remaining = usable_w - col_photo - col_dim - (col_price if show_price else 0)
    
    if show_advantage:
        col_name = remaining * 0.35
        col_adv = remaining * 0.65
    
    else:
        col_name = remaining
        col_adv = 0

    headers = [
        Paragraph("Photo",        styles["col_header"]),
        Paragraph("Product Name", styles["col_header"]),
        Paragraph("Dimension",   styles["col_header"]),
    ]
    col_widths = [col_photo, col_name, col_dim]

    if show_advantage:
        headers.append(Paragraph("Supply Advantage", styles["col_header"]))
        col_widths.append(col_adv)
    
    if show_price:
        headers.append(Paragraph("Price Range (THB/kg)", styles["col_header"]))
        col_widths.append(col_price)

    data = [headers]
    for p in products:
        row = [
            make_thumbnail(p["image"]),
            Paragraph(p["name"], styles["product_name"]),
            Paragraph(p["dim"],  styles["dim_text"]),
        ]
        if show_advantage:
            row.append(Paragraph(p["advantage"], styles["dim_text"]))
        if show_price:
            row.append(Paragraph(p["price"], styles["price_text"]))
        data.append(row)
        
    row_h = THUMB_SIZE + 4*mm

    t = Table(
        data,
        colWidths=col_widths,
        rowHeights=[10*mm] + [row_h] * len(products),
    )
    t.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), C_PRIMARY),
        ("ALIGN",          (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",         (0, 0), (-1, 0), "MIDDLE"),
        ("TOPPADDING",     (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING",  (0, 0), (-1, 0), 6),
        ("VALIGN",         (0, 1), (-1, -1), "MIDDLE"),
        ("ALIGN",          (0, 1), (0, -1), "CENTER"),
        ("LEFTPADDING",    (1, 1), (1, -1), 6),
        ("TOPPADDING",     (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING",  (0, 1), (-1, -1), 4),
        ("LINEBELOW",      (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
        ("LINEAFTER",      (0, 0), (-2, -1), 0.5, HexColor("#CCCCCC")),  # -2 = all cols except last
        ("LINEBEFORE",     (0, 0), (0, -1),  3,   C_ACCENT),
        ("BOX",            (0, 0), (-1, -1), 1,   C_PRIMARY),
    ]))
    return t


# ─────────────────────────────────────────────
#  PDF ASSEMBLER
# ─────────────────────────────────────────────

def generate_pdf(products, output_path):
    styles = make_styles()
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=30*mm, bottomMargin=20*mm,
        title="Product Suggestions Catalog",
        author=BRAND["company_name"],
        compress=1,
    )
    story = [
        Spacer(1, 4*mm),
        Paragraph("Product Suggestions Catalog", ParagraphStyle(
            "CatTitle", fontName="Helvetica-Bold",
            fontSize=18, textColor=C_PRIMARY, spaceAfter=12,
            leftIndent=-(7*mm),)),
        Paragraph(
            f"This catalog lists {len(products)} product(s) with specifications.",
            ParagraphStyle("Intro", fontName="Helvetica",
                fontSize=9, textColor=C_MID, spaceAfter=13,
                leftIndent=-(7*mm),)
        ),
        build_product_table(products, styles),
    ]
    doc.build(story, canvasmaker=BrandedCanvas)
    print(f"\n✅  PDF saved → {output_path}")


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    products = read_excel(PATHS["excel_file"])
    generate_pdf(products, PATHS["output_pdf"])